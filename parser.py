from openpyxl import load_workbook
import time
import logging
import os
from DrissionPage import ChromiumPage, ChromiumOptions ,errors
import argparse
import asyncio
import re
from app import send_document
parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('-F',type=str)
parser.add_argument('-B',type=str)
parser.add_argument('-O',type=str)
parser.add_argument('-Chat',type=str,default='')

# -F text.xlsx -B eqweq -Chat 545454 -O out.xlsx
args = parser.parse_args()
FILENAME = args.F +'.xlsx'
BOTKEY = args.B
OUTPUTFILENAME = args.O + '.xlsx'
CHAT_ID = args.Chat
LIMIT = 5
print(FILENAME,BOTKEY,OUTPUTFILENAME,CHAT_ID,LIMIT)

wt = load_workbook(FILENAME)
wsold = wt.worksheets[0]
wt.create_sheet('Парсер')
ws = wt['Парсер']

ws.append(['АУ Текст ','АУ Ссылка ','АУ ИНН','СРО АУ ','Должник Название','Ссылка на должника ','ИНН должника ','Тип сообщения','Ссылка на сообщение','Ссылка на бакрот базу','Адрес для корреспонденции АУ','E-mail АУ ','Вид торгов ','Дата публикации ','Дата торгов ','Сумма ','№ дела'])
def get_message(driver,message_link):
 
    feedtab= driver.new_tab(message_link)
    alltext = feedtab.ele('tag:body').text
    try:
          tabletorgi = feedtab.ele('tag:table@class=lotinfo',timeout=1.5)
    except Exception as e:
        print(e)
    try:
      tabletorgi = feedtab.ele('tag:table@class=lotInfo',timeout=1.5)
    except Exception as e:
        tabletorgi = None
      
    normalsum = 0
    if tabletorgi != None:
        if 'право требования' in tabletorgi.text or 'Права требования' in tabletorgi.text:
            return None, None ,None ,None ,None ,None,None
       
        if 'Отчет оценщика' in alltext:
                try:
                    tabletorgi =   feedtab.ele('tag:table@class=personInfo',2,1)
                except Exception as e:
                    pass
        if 'Начальная цена' in tabletorgi.text:
            
           
            sums = feedtab.eles('xpath://*[@id="ctl00_BodyPlaceHolder_lblBody"]/div/table[5]/tbody/tr/td[3]')
          
            for i in sums:
                txt = i.text.replace(' ','').replace(',','.')
                try:
                 normalsum+=float(txt)
                except:
                    pass
        if 'Начальная цена' and 'Шаг' in tabletorgi.text:
            
           
            sums = feedtab.eles('xpath://*[@id="ctl00_BodyPlaceHolder_lblBody"]/div/table[2]/tbody/tr/td[3]')
          
            for i in sums:
                txt = i.text.replace(' ','').replace(',','.')
                try:
                 normalsum+=float(txt)
                except:
                    pass
        if "Стоимость" in tabletorgi.text:
            
            sums = feedtab.eles('xpath://*[@id="ctl00_BodyPlaceHolder_lblBody"]/div/table[6]/tbody/tr/td[4]')
          
            for i in sums:
                txt = i.text.replace(' ','').replace(',','.')
                try:
                 normalsum+=float(txt)
                except:
                    pass
        if 'Лучшая цена,' in tabletorgi.text:
            sums = feedtab.eles('xpath://*[@id="ctl00_BodyPlaceHolder_lblBody"]/div/table[2]/tbody/tr/td[4]/div[2]')
          
            for i in sums:
                txt = i.text.replace(' ','').replace(',','.')
                try:
                 normalsum+=float(txt)
                except:
                    pass
        if 'торги признаны несостоявшимися' in tabletorgi.text:
                feedtab.close()
                return None, None ,None ,None ,None ,None,None
        
    try:
        normalsum = float(feedtab.ele('xpath://tr[./td[contains(text(),"Цена приобретения имущества")]]//td[2]',timeout=2).text.strip().replace(' ','').replace(',','.'))
        print(normalsum)
    except Exception as e:
        print(e)
      
    if normalsum < 5000000:
        feedtab.close()
        return None, None ,None ,None ,None ,None,None
    
    try:
       coreespondadress = feedtab.ele('xpath://tr[./td[contains(text(),"Адрес для корреспонденции")]]//td[2]',timeout=0.7).text
    except:
        coreespondadress = None
    print(coreespondadress)
    try:
       email = feedtab.ele('xpath://tr[./td[contains(text(),"E-mail")]]//td[2]',timeout=0.7).text
    except:
        email = None
    print(email)
    try:
       torgtype = feedtab.ele('xpath://tr[./td[contains(text(),"Вид торгов")]]//td[2]',timeout=0.7).text
    except:
        torgtype = None
    print(torgtype)
    try:
       publistiondate = feedtab.ele('xpath://tr[./td[contains(text(),"Дата публикации")]]//td[2]',timeout=0.7).text.strip()
    except:
       publistiondate = None
    print(publistiondate)
    if torgtype != None:
        try:
         dateandtorg =  feedtab.ele('xpath://tr[./td[contains(text(),"Дата и время торгов")]]//td[2]',timeout=0.7).text
        except Exception as e:
            dateandtorg = None
    else:
        dateandtorg = None
    dealnumber =  feedtab.ele('xpath://tr[./td[contains(text(),"№ дела")]]//td[2]',timeout=0.7).text
    feedtab.close()
    return coreespondadress,email,torgtype,publistiondate,dateandtorg,normalsum,dealnumber
 
      
   
        
    
    
def get_decorator(errors=(Exception, ), default_value=None):

    def decorator(func):

        def new_func(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except errors as e:
                print( "Got error! ", repr(e))
                return default_value

        return new_func

    return decorator
from concurrent.futures import ThreadPoolExecutor,wait

def get_bb_basa(bbtab,aruname):
     try:
       bbtab.ele('tag:input@id=email').input('sultanalievmaks1@mail.ru')
       bbtab.ele('tag:input@id=password').input('sultanalievmaks1')
       bbtab.ele('tag:button@type=submit').click()
     except Exception as e:
         print(e)
     try:
         bbtab.get('https://bankrotbaza.ru/reestr/arbitr')
         bbtab.ele('tag:input@placeholder=Поиск').input(aruname)
         hrefbbtab = bbtab.ele('tag:div@role=presentation').ele('tag:a').attr('href')
         bbtab.close()
         return hrefbbtab
     except Exception as e:
         bbtab.close()
         logging.error(e)
         return ' '
    


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('log.log', mode='a')
    ]
)
def get_user_inn(driver,url,name,msg,debtour_message_link):
    tab = driver.new_tab(url)
    inn =tab.ele('tag:span@id=ctl00_cphBody_lblINN').text
    tab.close()
    return inn,url,name,msg , debtour_message_link
def table_extractor(driver,classtable,listindextoreturn,message_link_bool):
    
    table = driver.ele(f'tag:table@class={classtable}',timeout=1)
    trlist = table.eles('tag:tr',timeout=0.5)
    tablelist = []
   
    for i in trlist:
      try:    
           tdlist = i.eles('tag:td',timeout=0.5)
           if tdlist != []:
               templist = []
               for x in listindextoreturn:
                    templist.append(tdlist[x].text )
                   
                       
                       
                    
               for x in listindextoreturn:
                     if x == 0 or x ==2:
                      
                     
                        href = tdlist[x].ele('tag:a',timeout=0.5).attr('href')
                        templist.append(href)
                     if x==1 and message_link_bool== True:
                         onlcik = tdlist[x].ele('tag:a',timeout=0.5).attr('onclick')
                         messagelink = re.findall("openNewWin\('([^']*)'",onlcik)[0]
                         
                         templist.append(messagelink)
                         
                     
               tablelist.append(templist)
                 
                    
                    
                
      except Exception as e:
        
          logging.error(e)
          continue
           
    return tablelist


         
          

def get_chromium_options(browser_path: str, arguments: list) -> ChromiumOptions:
    """
    Configures and returns Chromium options.
    
    :param browser_path: Path to the Chromium browser executable.
    :param arguments: List of arguments for the Chromium browser.
    :return: Configured ChromiumOptions instance.
    """
    options = ChromiumOptions()
    options.set_paths(browser_path=browser_path)
    options.set_load_mode('eager')
    options.headless(False)
    

    for argument in arguments:
        options.set_argument(argument)
    return options


    

def main(wsold,wsnew,wt):
    counter = 0
    # Chromium Browser Path
    BASE_LINK = 'https://old.bankrot.fedresurs.ru'
   
    browser_path = os.getenv('CHROME_PATH', os.path.abspath(os.getcwd()))
    # Windows Example
    # browser_path = os.getenv('CHROME_PATH', r"C:/Program Files/Google/Chrome/Application/chrome.exe")
    message_list = ["Отчет оценщика об оценке имущества должника","Сведения о заключении договора купли-продажи","Объявление о проведении торгов","Сообщение о результатах торгов"]
    for i in range(len(message_list)):
        message_list[i] = message_list[i].replace(' ','').strip().lower()
   
    # Arguments to make the browser better for automation and less detectable.
    arguments = [
        "-no-first-run",
        # "-force-color-profile=srgb",
        "-metrics-recording-only",
        "-password-store=basic",
        "-use-mock-keychain",
        "-export-tagged-pdf",
        "-no-default-browser-check",
        "-disable-background-mode",
        "-enable-features=NetworkService,NetworkServiceInProcess,LoadCryptoTokenExtension,PermuteTLSExtensions",
        "-disable-features=FlashDeprecationWarning,EnablePasswordsAccountStorage",
        "-deny-permission-prompts",
      #  "-remote-debugging-pipe",
         #'-disable-dev-shm-usage'
        
       
        "-accept-lang=en-US",
     
       '--no-sandbox'
      
    ]
    nosearches = 'По заданным критериям не найдено ни одной записи. Уточните критерии поиска.'
    options = get_chromium_options(browser_path, arguments)
    options.set_user_agent('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36')
    
 
    options.set_user_data_path(r'./tmp'+OUTPUTFILENAME)
    
    
    
    options.set_pref('webrtc.ip_handling_polic','disable_non_proxied_udp')
    options.set_pref('webrtc.multiple_routes_enabled',False)
    options.set_pref('webrtc.nonproxied_udp_enabled',False)
   
    options.no_imgs()
    port = 9226
    options.set_local_port(port)
   

   
    # Initialize the browser
   
   
   
    driver = ChromiumPage(addr_or_opts=options)
   
   
    
   
   
    try:
          for i in range(wsold.min_row+1 , wsold.max_row+1):
                aruname,arulink,arunINN,aruSRO,debtorname,debtor_link,debtor_INN =[wsold.cell(i,x).value for x in range(1,8)]
                print(aruname)
                debtor_message_link = None
                bblinktab = None
                start = time.perf_counter()
                if aruname != '' and aruname != None:
                            bbtab= driver.new_tab('https://bankrotbaza.ru/login')
                            with ThreadPoolExecutor(1) as executor:
                                feature = executor.submit(get_bb_basa,bbtab,aruname)
                                bblinktab = feature.result()
                                
                                
                            
                            driver.get('https://old.bankrot.fedresurs.ru/ArbitrManagersList.aspx')
                            lastnameinput= driver.ele('tag:input@@name=ctl00$cphBody$ArbitrManagerList1$tbLastName')
                            firstnameinput=driver.ele('tag:input@@name=ctl00$cphBody$ArbitrManagerList1$tbFirstName')
                            middlenameinput =driver.ele('tag:input@@name=ctl00$cphBody$ArbitrManagerList1$tbMiddleName')
                            findelem = driver.ele('tag:input@@name=ctl00$cphBody$ArbitrManagerList1$ibArmSearch')
                            aruname = aruname.strip().split(' ')
                            lastnameinput.clear()
                            firstnameinput.clear()
                            middlenameinput.clear()
                            lastnameinput.input(aruname[0])
                            firstnameinput.input(aruname[1])
                            middlenameinput.input(aruname[2])
                            findelem.click()
                            driver.wait(1)
                           
                            table  = driver.ele('tag:table').text
                            if nosearches in table:
                                continue
                            for i in table_extractor(driver,'bank',[0,3],False):
                                arutext = i[0]
                                aruname = arutext
                                aruSRO = i[1]
                                arulink = i[2]
                                
                                
                                driver.get(arulink)
                                lastpage = 1
                                clickelem = driver.ele('xpath://td/a[contains(@href,"Page$1")]')
                               
                                while clickelem or lastpage!= 0:
                                    driver.wait(1)
                                    if lastpage == LIMIT:
                                        break
                                    arunINN = driver.ele('tag:tr@id=ctl00_cphBody_trInn').eles('tag:td')[1].text.strip()
                                  
                                    tasks = []
                                   
                                  
                                    with ThreadPoolExecutor(3) as executor:
                                        for j in table_extractor(driver,'bank',[1,2],True):
                                            debrourmessage = j[0].strip().replace(' ','').lower()
                                            result = []
                                            
                                            if not 'Page' in j[2] and   debrourmessage in message_list:
                                                    debtorname = j[1].strip()
                                                    debrourmessage= j[0].strip()
                                                    
                                                    debtor_link = j[3].strip()
                                                   
                                                    try:
                                                        debtor_message_link =BASE_LINK+ j[2 ].strip()
                                                    except Exception as e:
                                                        logging.error('Messagelink error'+str(e))
                                               #     print(debtorname,debrourmessage,debtor_link,debtor_message_link)
                                                    task = executor.submit(get_user_inn,driver,debtor_link,debtorname,debrourmessage,debtor_message_link)
                                                    tasks.append(task)
                                        futures, _ = wait(tasks)
                                        with ThreadPoolExecutor(3) as executor:
                                            for future in futures:
                                                debtor_INN,debtor_link,debtorname,debrourmessage,debtor_message_link = future.result()
                                                fs = executor.submit(get_message,driver,debtor_message_link)
                                                coreespondadress,email,torgtype,publistiondate,dateandtorg,normalsum,dealnumber = fs.result()
                                               # print( coreespondadress,email,torgtype,publistiondate,dateandtorg,normalsum,dealnumber)
                                              
                                      
                                         
                                                ws.append([ aruname,arulink,arunINN,aruSRO,debtorname,debtor_link,debtor_INN,debrourmessage,debtor_message_link,bblinktab
                                                           , coreespondadress,email,torgtype,publistiondate,dateandtorg,normalsum,dealnumber])
                                       
                                    try:
                                        clickelem =driver.ele(f'xpath://td/a[contains(@href,"Page${lastpage+1}")]')
                                        clickelem.click()
                                       
                                        
                                        if clickelem.text.strip() == '...':
                                            lastpage+=1
                                        else:
                                            lastpage = int(clickelem.text.strip())
                                     
                                    except:
                                        lastpage=0
                wt.save(OUTPUTFILENAME)
                stop = time.perf_counter() - start
                print(stop)                                                                    
                                
                             
                                    
                                        
                                    
                                 
                                    
                                    
                            
                                
                                
                                
                                
                            

                            
                            
                            
                            
    except Exception as e:
        logging.error("An error occurred: %s", str(e))
        logging.exception(e)
        
    finally:
        driver.get_screenshot('./tmp','2.jpg')
        logging.info('Closing the browser.')
        driver.quit()
    

            
if __name__ == '__main__':
    try:
     main(wsold,ws,wt)  
    except Exception as e:
        
        print(e)
        
    asyncio.run(send_document(CHAT_ID,OUTPUTFILENAME))
           
           